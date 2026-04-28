# test_new_handlers.py - smoke test for new format handlers
import os, sys, traceback

basedir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, basedir)
os.chdir(basedir)

FIXTURES = os.path.join(basedir, 'tests', 'fixtures')

def create_odt(path):
    import zipfile, tempfile, shutil
    tmp = tempfile.mkdtemp()
    os.makedirs(os.path.join(tmp, 'META-INF'))
    with open(os.path.join(tmp, 'mimetype'), 'w', encoding='utf-8') as f:
        f.write('application/vnd.oasis.opendocument.text')
    with open(os.path.join(tmp, 'meta.xml'), 'w', encoding='utf-8') as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?><office:document-meta xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" xmlns:meta="urn:oasis:names:tc:opendocument:xmlns:meta:1.0"><office:meta><meta:generator>SM</meta:generator></office:meta></office:document-meta>')
    with open(os.path.join(tmp, 'content.xml'), 'w', encoding='utf-8') as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?><office:document-content xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"><office:body><office:text><text:p>Test</text:p></office:text></office:body></office:document-content>')
    with open(os.path.join(tmp, 'META-INF', 'manifest.xml'), 'w', encoding='utf-8') as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?><manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0" manifest:version="1.2"><manifest:file-entry manifest:media-type="application/vnd.oasis.opendocument.text" manifest:full-path="/"/></manifest:manifest>')
    with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(os.path.join(tmp, 'mimetype'), 'mimetype', compress_type=zipfile.ZIP_STORED)
        for root_dir, dirs, files in os.walk(tmp):
            for file in files:
                full = os.path.join(root_dir, file)
                arcname = os.path.relpath(full, tmp).replace(os.sep, '/')
                if arcname != 'mimetype':
                    zf.write(full, arcname)
    shutil.rmtree(tmp)

def create_xlsx(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Hello'
    for i in range(2, 50):
        ws[f'A{i}'] = f'Item {i}'
    wb.save(path)

def create_rtf(path):
    with open(path, 'w', encoding='utf-8') as f:
        f.write(r'{\rtf1\ansi\deff0 {\fonttbl{\f0 Arial;}} \f0\fs24 Test RTF Document }')

def create_epub(path):
    import zipfile, tempfile, shutil
    tmp = tempfile.mkdtemp()
    os.makedirs(os.path.join(tmp, 'META-INF'))
    with open(os.path.join(tmp, 'mimetype'), 'w', encoding='utf-8') as f:
        f.write('application/epub+zip')
    with open(os.path.join(tmp, 'META-INF', 'container.xml'), 'w', encoding='utf-8') as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?><container version="1.0" xmlns="urn:oasis:names:tc:opendocument:xmlns:container"><rootfile full-path="OEBPS/content.opf" media-type="application/oebps-package+xml"/></container>')
    os.makedirs(os.path.join(tmp, 'OEBPS'))
    with open(os.path.join(tmp, 'OEBPS', 'content.opf'), 'w', encoding='utf-8') as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?><package xmlns="http://www.idpf.org/2007/opf" version="2.0"><metadata xmlns:dc="http://purl.org/dc/elements/1.1/"><dc:title>Test</dc:title></metadata><manifest/><spine></spine></package>')
    with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(os.path.join(tmp, 'mimetype'), 'mimetype', compress_type=zipfile.ZIP_STORED)
        for root_dir, dirs, files in os.walk(tmp):
            for file in files:
                full = os.path.join(root_dir, file)
                arcname = os.path.relpath(full, tmp).replace(os.sep, '/')
                if arcname != 'mimetype':
                    zf.write(full, arcname)
    shutil.rmtree(tmp)

from src.core.manager import StealthMark
from PIL import Image

wm = StealthMark()
MARK = 'SM-2026'
tests = []

# Document formats
doc_tests = [
    ('XLSX', 'test.xlsx', create_xlsx),
    ('ODT', 'test.odt', create_odt),
    ('ODS', 'test.ods', create_odt),
    ('RTF', 'test.rtf', create_rtf),
    ('EPUB', 'test.epub', create_epub),
]

for name, fname, creator in doc_tests:
    path = os.path.join(FIXTURES, fname)
    try:
        creator(path)
        stem, ext = os.path.splitext(fname)
        out = os.path.join(FIXTURES, f'{stem}_out{ext}')
        r1 = wm.embed(path, MARK, out)
        r2 = wm.extract(out)
        ok = r2.is_success and r2.watermark and r2.watermark.content == MARK
        tests.append((name, ok, '' if ok else f'{r1.status}/{r2.message}'))
    except Exception as e:
        tests.append((name, False, str(e)[:60]))

# Image formats
for ext in ['tiff', 'webp', 'gif']:
    name = ext.upper()
    src_path = os.path.join(FIXTURES, f'test.{ext}')
    out_path = os.path.join(FIXTURES, f'out_{ext}.{ext}')
    try:
        img = Image.new('RGB', (200, 200), color='red')
        img.save(src_path, format=ext.upper() if ext != 'gif' else 'GIF')
        r1 = wm.embed(src_path, MARK, out_path)
        r2 = wm.extract(out_path)
        ok = r2.is_success and r2.watermark and r2.watermark.content == MARK
        tests.append((name, ok, '' if ok else f'{r1.status}/{r2.message}'))
    except Exception as e:
        tests.append((name, False, str(e)[:60]))

print('=== New Format Embed+Extract Smoke Test ===')
for name, ok, err in tests:
    status = '[OK]' if ok else '[FAIL]'
    detail = '' if ok else f' -> {err}'
    print(f'  {status} {name}{detail}')
passed = sum(1 for _,ok,_ in tests if ok)
print(f'Result: {passed}/{len(tests)} passed')
